#!/usr/bin/env python3
"""
금융 RFP 빠른 테스트 — PDF 변환 생략(extract-only), 병렬 샘플, ETA 표시.

예:
  python scripts/run_financial_test.py
  python scripts/run_financial_test.py --only 001-하나은행
  python scripts/run_financial_test.py --workers 2 --llm-concurrency 24
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

from tqdm import tqdm

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
FINAL_MANIFEST = ROOT / "data" / "final.manifest.json"
DOMAIN_TEST = ROOT / "data" / "processed" / "domain_test"

sys.path.insert(0, str(BACKEND))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("financial_test")

FINANCIAL_IDS = frozenset({"001-하나은행", "002-신한라이프"})
DOMAIN_NAMES = {"001-하나은행": "하나.xlsx", "002-신한라이프": "신한.xlsx"}

# extract-only + prefetch 병렬 기준 경험치(초/샘플)
_EST_SEC = {"001-하나은행": 150, "002-신한라이프": 240}


def _estimate_seconds(samples: list[dict], workers: int) -> int:
    per = [_EST_SEC.get(s["id"], 180) for s in samples]
    if workers <= 1:
        return sum(per)
    # 병렬: 워커당 큐 길이 합의 최대
    buckets = [0] * workers
    for t in sorted(per, reverse=True):
        buckets[buckets.index(min(buckets))] += t
    return max(buckets)


def _validate_xlsx(path: Path) -> list[str]:
    """간단 품질 검사 — 오류 메시지 목록."""
    import openpyxl

    errs: list[str] = []
    if not path.is_file():
        return [f"파일 없음: {path}"]
    wb = openpyxl.load_workbook(path, read_only=True)
    if not wb.sheetnames:
        errs.append("시트 없음")
    for sn in wb.sheetnames:
        if sn == "개요":
            continue
        ws = wb[sn]
        if ws.max_column and ws.max_column < 4:
            errs.append(f"[{sn}] 칼럼 수 부족 (카테고리 칼럼 누락?)")
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if hdr and "카테고리" not in (hdr[1] if len(hdr) > 1 else ""):
            if "카테고리" not in hdr:
                errs.append(f"[{sn}] '카테고리' 헤더 없음: {hdr[:4]}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            group = str(row[0] or "")
            if " — " in group and len(group) > 40:
                errs.append(f"[{sn}] 요건구분에 계위 병합: {group[:50]}…")
                break
            if group.startswith("2.3.2. 기타") and "본 프로젝트" in group:
                errs.append(f"[{sn}] 기타 요구사항 서술형이 요건구분에 포함")
                break
    wb.close()
    return errs


def _run_one(
    entry: dict,
    out_root: Path,
    *,
    extract_only: bool,
    llm_concurrency: int,
) -> dict:
    os.environ["LLM_CONCURRENCY"] = str(llm_concurrency)
    from app.services.pipeline_logger import PipelineLogSession, set_active_session
    from prototype.v3.pipeline_final import run_sample

    sample_id = entry["id"]
    source = ROOT / entry["source"]
    gold = ROOT / entry["gold_xlsx"] if entry.get("gold_xlsx") else None
    t0 = time.time()
    out: dict = {"id": sample_id, "elapsed_sec": 0.0}

    session = PipelineLogSession(
        out_root / sample_id,
        run_id="logs",
        source_path=source,
        source_name=source.name,
        engine="v3-financial-test",
    )
    set_active_session(session)
    try:
        manifest = run_sample(
            sample_id=sample_id,
            source=source,
            strategy=entry["strategy"],
            label=entry.get("label", sample_id),
            out_root=out_root,
            gold_xlsx=gold if gold.is_file() else None,
            log_session=session,
            extract_only=extract_only,
            llm_concurrency=llm_concurrency,
        )
        out["report"] = manifest.get("report", {})
        out["steps"] = manifest.get("steps", [])[-2:]
        xlsx = out_root / sample_id / "requirements.xlsx"
        DOMAIN_TEST.mkdir(parents=True, exist_ok=True)
        dest = DOMAIN_TEST / DOMAIN_NAMES.get(sample_id, f"{sample_id}.xlsx")
        if xlsx.is_file():
            dest.write_bytes(xlsx.read_bytes())
            out["domain_test"] = str(dest)
            out["validation"] = _validate_xlsx(dest)
    except Exception as e:
        out["error"] = str(e)
        out["trace"] = traceback.format_exc()[-800:]
    finally:
        set_active_session(None)
        out["elapsed_sec"] = round(time.time() - t0, 1)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="금융 RFP 빠른 테스트 (extract-only + 병렬)")
    parser.add_argument("--only", default="001-하나은행,002-신한라이프", help="sample id 쉼표 구분")
    parser.add_argument("--workers", type=int, default=2, help="샘플 병렬 수")
    parser.add_argument("--llm-concurrency", type=int, default=20, help="샘플 내 LLM 동시 호출")
    parser.add_argument("--full", action="store_true", help="PDF 변환 포함 (느림)")
    args = parser.parse_args()

    meta = json.loads(FINAL_MANIFEST.read_text(encoding="utf-8"))
    out_root = ROOT / meta.get("artifacts_root", "data/artifacts_final")
    allow = {s.strip() for s in args.only.split(",") if s.strip()}
    samples = [s for s in meta["samples"] if s["id"] in allow and s["id"] in FINANCIAL_IDS]
    if not samples:
        log.error("금융 샘플 없음: %s", args.only)
        return 1

    extract_only = not args.full
    est = _estimate_seconds(samples, args.workers)
    eta = datetime.now() + timedelta(seconds=est)
    print(
        f"\n금융 테스트 시작 — {len(samples)}개 샘플\n"
        f"  모드: {'extract-only (HTML 캐시)' if extract_only else 'full (PDF 변환 포함)'}\n"
        f"  workers={args.workers}  llm_concurrency={args.llm_concurrency}\n"
        f"  예상 소요: ~{est // 60}분 {est % 60}초\n"
        f"  예상 완료: {eta.strftime('%H:%M:%S')}\n"
    )

    t0 = time.time()
    results: list[dict] = []

    if args.workers <= 1:
        for entry in tqdm(samples, desc="financial"):
            results.append(
                _run_one(
                    entry,
                    out_root,
                    extract_only=extract_only,
                    llm_concurrency=args.llm_concurrency,
                )
            )
    else:
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futs = {
                pool.submit(
                    _run_one,
                    e,
                    out_root,
                    extract_only=extract_only,
                    llm_concurrency=args.llm_concurrency,
                ): e["id"]
                for e in samples
            }
            for fut in tqdm(as_completed(futs), total=len(futs), desc="financial"):
                results.append(fut.result())

    elapsed = time.time() - t0
    ok = [r for r in results if "error" not in r]
    fail = [r for r in results if "error" in r]

    print(f"\n=== 완료 {len(ok)}/{len(results)} — 실제 {elapsed:.0f}초 (예상 {est}초) ===")
    for r in sorted(results, key=lambda x: x["id"]):
        if "error" in r:
            print(f"  ✗ {r['id']}: {r['error']}")
            if r.get("trace"):
                print(f"      {r['trace'][:200]}")
        else:
            rep = r.get("report", {})
            val = r.get("validation") or []
            flag = f" ⚠ {len(val)}건" if val else ""
            gold = rep.get("gold_recall")
            gold_s = f", gold {gold['recall']*100:.1f}% ({gold['covered']}/{gold['gold_total']})" if gold else ""
            print(
                f"  ✓ {r['id']}: {rep.get('extracted_rows')} rows, "
                f"{r['elapsed_sec']}s{gold_s} → {r.get('domain_test', '')}{flag}"
            )
            for v in val[:3]:
                print(f"      - {v}")
    if fail:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
