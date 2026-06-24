#!/usr/bin/env python3
"""
패키지 샘플 다층 inspect — HTML/추출/로직/LLM 레이어별 진단 + gold recall.

출력: temp/inspect_report.md, temp/inspect_summary.json
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from tqdm import tqdm

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
PACKAGES = ROOT / "data" / "packages"
RAW = ROOT / "data" / "raw"
OUT = ROOT / "temp"
sys.path.insert(0, str(BACKEND))

GOLD_MAP = {
    "003-신한라이프": "003-신한라이프-gold.xlsx",
    "004-JB금융": "004-JB금융-gold.xlsx",
    "005-법제처": "005-법제처-gold.xlsx",
}


def _analyze_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True)
    stats = {
        "tabs": 0,
        "rows": 0,
        "empty_both": 0,
        "hw_noise": 0,
        "prefixes": Counter(),
    }
    for ws in wb.worksheets:
        if ws.title == "개요":
            continue
        stats["tabs"] += 1
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < 4 or not r[3]:
                continue
            stats["rows"] += 1
            if not r[1] and not r[2]:
                stats["empty_both"] += 1
            d = str(r[3])
            if "K8S Worker" in d or (d.startswith("□") and "CPU" in d):
                stats["hw_noise"] += 1
            rid = str(r[0] or "")
            if "_" in rid:
                stats["prefixes"][rid.rsplit("_", 1)[0]] += 1
    wb.close()
    stats["prefix_top"] = stats["prefixes"].most_common(6)
    del stats["prefixes"]
    return stats


def _recall(out: Path, gold: Path) -> dict | None:
    from prototype.v2.validate import completeness, load_gold
    from prototype.v2.extract import Req

    if not gold.is_file():
        return None
    gold_items = load_gold(str(gold))
    wb = openpyxl.load_workbook(out, read_only=True)
    reqs = []
    for ws in wb.worksheets:
        if ws.title == "개요":
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) < 4 or not r[3]:
                continue
            reqs.append(
                Req(
                    doc="i",
                    table_id=0,
                    page=1,
                    tab=ws.title,
                    top=str(r[1] or ""),
                    mid=str(r[2] or ""),
                    detail=str(r[3]),
                )
            )
    wb.close()
    res = completeness(gold_items, reqs)
    return {
        "recall": round(res["recall"], 4),
        "missing": len(res["missing"]),
        "gold_total": res["gold_total"],
        "missing_by_sheet": res["missing_by_sheet"],
    }


def _inspect_pdf_layers(pdf: Path) -> dict:
    """HTML(JSON) + 추출 후보 레이어."""
    import json as _json
    from prototype.v2.document import extract_document

    layer: dict = {"kind": "pdf"}
    try:
        import opendataloader_pdf
        work = OUT / "inspect_convert" / pdf.stem
        work.mkdir(parents=True, exist_ok=True)
        for stale in work.glob("*.json"):
            stale.unlink()
        opendataloader_pdf.convert(input_path=[str(pdf)], output_dir=str(work), format="json")
        jp = next(work.glob("*.json"))
        doc = _json.loads(jp.read_text(encoding="utf-8"))

        def count_types(nodes, acc=None):
            acc = acc or Counter()
            for el in nodes:
                if not isinstance(el, dict):
                    continue
                acc[el.get("type", "?")] += 1
                kids = el.get("kids")
                if isinstance(kids, list):
                    count_types(kids, acc)
            return acc

        types = count_types(doc.get("kids", []))
        layer["doc_types"] = dict(types)
        _, cands = extract_document(pdf.stem, doc, "fine", defer_tables=True)
        ncol = Counter(g.ncols for g, _ in cands)
        list_blocks = sum(1 for g, _ in cands if g.ncols == 1 and g.table_id < 0)
        layer["candidates"] = len(cands)
        layer["list_blocks"] = list_blocks
        layer["ncol_dist"] = dict(ncol)
    except Exception as e:
        layer["error"] = str(e)
    return layer


def _run_one(pkg_base: str, src: Path) -> dict:
    from prototype.v2.pipeline import run as v2_run

    ext = src.suffix.lower()
    korean = ext in (".hwpx", ".hwp", ".html", ".htm")
    gold_name = GOLD_MAP.get(pkg_base)
    gold = PACKAGES / gold_name if gold_name else None

    report: dict = {"package": pkg_base, "source": str(src)}
    if ext == ".pdf" and src.is_file():
        report["layers"] = {"extract": _inspect_pdf_layers(src)}

    manifest = v2_run(str(src), str(gold) if gold and gold.is_file() else None, mode="llm", korean_hwp=korean)
    xlsx = Path(manifest["artifacts"]["requirements_xlsx"])
    report["steps"] = manifest["steps"]
    report["rows"] = manifest["report"]["extracted_rows"]
    report["output"] = {"xlsx": str(xlsx), "stats": _analyze_xlsx(xlsx)}
    if gold and gold.is_file():
        report["recall"] = _recall(xlsx, gold)
    if "recall" in manifest.get("report", {}):
        report["pipeline_recall"] = manifest["report"]["recall"]

    # 로직/LLM 이슈 요약
    issues: list[str] = []
    st = report["output"]["stats"]
    if st["hw_noise"]:
        issues.append(f"HW노이즈 {st['hw_noise']}행 잔존")
    if st["rows"] and st["empty_both"] / st["rows"] < 0.25:
        issues.append(f"병합형 빈칸 비율 낮음 ({st['empty_both']}/{st['rows']})")
    for step in manifest["steps"]:
        if "정합성 검토" in step:
            issues.append(f"정합성: {step}")
    if report.get("recall") and report["recall"]["recall"] < 0.9:
        issues.append(f"recall {report['recall']['recall']*100:.1f}%")
    report["issues"] = issues
    return report


def _markdown(reports: list[dict]) -> str:
    lines = [
        "# 패키지 Inspect 리포트",
        f"\n생성: {datetime.now().isoformat(timespec='seconds')}\n",
        "| 패키지 | 행 | 탭 | 빈칸 | recall | 이슈 |",
        "|--------|-----|-----|------|--------|------|",
    ]
    for r in reports:
        if "error" in r:
            lines.append(f"| {r['package']} | — | — | — | — | {r['error']} |")
            continue
        st = r["output"]["stats"]
        rec = r.get("recall", {})
        rc = f"{rec['recall']*100:.1f}%" if rec else "—"
        iss = "; ".join(r.get("issues", [])) or "—"
        lines.append(
            f"| {r['package']} | {st['rows']} | {st['tabs']} | {st['empty_both']} | {rc} | {iss} |"
        )
    lines.append("\n## 레이어별 상세\n")
    for r in reports:
        if "error" in r:
            continue
        lines.append(f"### {r['package']}\n")
        if "layers" in r:
            ex = r["layers"].get("extract", {})
            if ex:
                lines.append(f"- **HTML/추출**: types={ex.get('doc_types')} candidates={ex.get('candidates')} list_blocks={ex.get('list_blocks')}")
        lines.append(f"- **파이프라인**: {r['rows']}행")
        for s in r.get("steps", []):
            if any(k in s for k in ("extract", "탭", "비요구", "LLM", "gold", "리스트", "정합성")):
                lines.append(f"  - {s}")
        if r.get("issues"):
            lines.append(f"- **이슈**: {', '.join(r['issues'])}")
        lines.append("")
    return "\n".join(lines)


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    pkgs = sorted(PACKAGES.glob("*.*"))
    bases = sorted({p.stem.rsplit("-gold", 1)[0] for p in pkgs if p.suffix.lower() != ".xlsx" or "-gold" not in p.stem})
    # unique package bases from pdf/doc/hwpx
    sources = []
    for p in sorted(PACKAGES.iterdir()):
        if p.suffix.lower() in (".pdf", ".doc", ".docx", ".hwp", ".hwpx", ".html", ".htm"):
            if "-gold" not in p.stem:
                sources.append(p)

    reports = []
    for src in tqdm(sources, desc="inspect"):
        base = src.stem
        try:
            reports.append(_run_one(base, src))
        except Exception as e:
            reports.append({"package": base, "error": str(e)})

    summary_path = OUT / "inspect_summary.json"
    summary_path.write_text(json.dumps(reports, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path = OUT / "inspect_report.md"
    md_path.write_text(_markdown(reports), encoding="utf-8")
    print(f"\n리포트: {md_path}\n요약: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
