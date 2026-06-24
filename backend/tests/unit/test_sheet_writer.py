from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.domain.enums import ExportMode, Judgement
from app.domain.models import HumanJudgement, Recommendation, Requirement
from app.phase1.writers.export_columns import EXPORT_COLUMNS, EXPORT_PRESETS, resolve_export_columns
from app.phase1.writers.sheet_writer import RequirementSheetWriter


def _headers_for(reqs: list[Requirement], keys: list[str] | None = None) -> list[str]:
    col_keys = resolve_export_columns(reqs, {}, {}, ExportMode.BOTH, keys)
    return [EXPORT_COLUMNS[k].header for k in col_keys]


def _req(rid: str, cat: str, code: str, detail: str, **kwargs) -> Requirement:
    return Requirement(
        id=rid, doc_id="d", category=cat, code=code, name=detail[:20], detail=detail, **kwargs
    )


def test_writer_produces_summary_and_category_sheets(tmp_path: Path) -> None:
    reqs = [
        _req("r1", "데이터 수집", "DATA-001", "원천 시스템 연계"),
        _req("r2", "데이터 수집", "DATA-002", "지원 파일 형식"),
        _req("r3", "저장 구조", "STOR-001", "Object Storage"),
    ]
    out = RequirementSheetWriter().write(tmp_path / "out.xlsx", reqs)
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "개요"
    assert "데이터 수집" in wb.sheetnames
    assert "저장 구조" in wb.sheetnames

    detail = wb["데이터 수집"]
    headers = list(next(detail.iter_rows(values_only=True)))
    assert "항목명" in headers
    assert "상세요건" in headers


def _blank(v: object) -> bool:
    return v is None or v == ""


def test_mode_ai_only_blanks_human_cols(tmp_path: Path) -> None:
    reqs = [_req("r1", "분류A", "A-001", "본문")]
    recs = {
        "r1": Recommendation(
            requirement_id="r1",
            ai_risk=Judgement.PARTIAL,
            ai_reason="부분 가능",
            missing_tech=["X"],
            consortium_need="Y",
        )
    }
    judges = {"r1": HumanJudgement(requirement_id="r1", mark=Judgement.YES, note="OK")}

    out = RequirementSheetWriter().write(
        tmp_path / "ai.xlsx",
        reqs,
        recommendations=recs,
        judgements=judges,
        mode=ExportMode.AI,
        columns=EXPORT_PRESETS["full"],
    )
    wb = load_workbook(out)
    ws = wb["분류A"]
    headers = list(next(ws.iter_rows(values_only=True)))
    body = list(ws.iter_rows(values_only=True))[1]
    row = dict(zip(headers, body, strict=True))
    assert row["AI 리스크"] == "△"
    assert row["AI 이유"] == "부분 가능"
    assert _blank(row.get("사람 판정"))
    assert _blank(row.get("사람 메모"))


def test_ordered_layout_splits_by_category_tab(tmp_path: Path) -> None:
    reqs = [
        _req("r1", "ICT 인프라 요구사항", "A-001", "첫 탭 1", source_page=2),
        _req("r2", "ICT 인프라 요구사항", "A-002", "첫 탭 2", source_page=2),
        _req("r3", "정보보호 요청사항", "B-001", "둘째 탭", source_page=5),
    ]
    out = RequirementSheetWriter().write(tmp_path / "ordered.xlsx", reqs, layout="ordered")
    wb = load_workbook(out)
    data_sheets = [s for s in wb.sheetnames if s != "개요"]
    assert len(data_sheets) == 2
    assert "ICT 인프라 요구사항" in data_sheets
    assert "정보보호 요청사항" in data_sheets


def test_mode_human_only_blanks_ai_cols(tmp_path: Path) -> None:
    reqs = [_req("r1", "분류A", "A-001", "본문")]
    judges = {"r1": HumanJudgement(requirement_id="r1", mark=Judgement.NO, note="제외")}

    out = RequirementSheetWriter().write(
        tmp_path / "h.xlsx",
        reqs,
        judgements=judges,
        mode=ExportMode.HUMAN,
        columns=EXPORT_PRESETS["full"],
    )
    wb = load_workbook(out)
    ws = wb["분류A"]
    headers = list(next(ws.iter_rows(values_only=True)))
    body = list(ws.iter_rows(values_only=True))[1]
    row = dict(zip(headers, body, strict=True))
    assert _blank(row.get("AI 리스크"))
    assert _blank(row.get("AI 이유"))
    assert row["사람 판정"] == "X"
    assert row["사람 메모"] == "제외"
