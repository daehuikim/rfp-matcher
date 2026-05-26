from __future__ import annotations

import io
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.domain.enums import DocumentMime, Judgement
from app.domain.models import Document, HumanJudgement, Recommendation, Requirement
from app.main import create_app
from app.phase1.writers.sheet_writer import COL_HEADERS


@pytest.fixture
def seeded_client(tmp_path: Path) -> TestClient:
    app = create_app()
    client = TestClient(app)
    client.__enter__()  # lifespan 트리거
    container = app.state.container
    container.settings.storage_root = tmp_path

    # repo에 직접 시드 — 업로드 파이프라인 우회
    import asyncio

    async def seed() -> None:
        doc = Document(
            id="doc-1",
            src_path=tmp_path / "sample.pdf",
            mime=DocumentMime.PDF,
        )
        await container.repo.save_document(doc)
        reqs = [
            Requirement(
                id="r1",
                doc_id="doc-1",
                category="데이터",
                code="DAT-001",
                name="원천 시스템 연계",
                detail="다양한 원천 시스템에서 데이터 수집",
            ),
            Requirement(
                id="r2",
                doc_id="doc-1",
                category="저장",
                code="STO-001",
                name="Object Storage 구성",
                detail="Object Storage 구성 전략",
            ),
        ]
        await container.repo.save_requirements("doc-1", reqs)
        await container.repo.upsert_recommendation(
            Recommendation(
                requirement_id="r1",
                ai_risk=Judgement.YES,
                ai_reason="카탈로그로 커버 가능",
                missing_tech=[],
                consortium_need=None,
            )
        )
        await container.repo.upsert_judgement(
            HumanJudgement(requirement_id="r2", mark=Judgement.PARTIAL, note="검토 중")
        )

    asyncio.run(seed())
    yield client
    client.__exit__(None, None, None)


def _open_first_data_sheet(content: bytes):
    wb = load_workbook(io.BytesIO(content))
    # 총괄표 다음 첫 분류 시트를 데이터 시트로 사용 — 분류명 정렬 순서
    cat_sheets = [s for s in wb.sheetnames if s != "총괄표"]
    return wb[cat_sheets[0]]


def test_export_mode_ai_only_blanks_human(seeded_client: TestClient) -> None:
    r = seeded_client.get("/documents/doc-1/export?mode=ai")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd")
    ws = _open_first_data_sheet(r.content)
    # 헤더 확인
    headers = next(ws.iter_rows(values_only=True))
    assert list(headers) == COL_HEADERS
    # 모든 데이터 행에서 사람 컬럼 비어 있어야 함
    for row in list(ws.iter_rows(values_only=True))[1:]:
        d = dict(zip(COL_HEADERS, row, strict=True))
        assert d["사람 판정"] in (None, "")
        assert d["사람 메모"] in (None, "")


def test_export_mode_human_only_blanks_ai(seeded_client: TestClient) -> None:
    r = seeded_client.get("/documents/doc-1/export?mode=human")
    assert r.status_code == 200
    ws = _open_first_data_sheet(r.content)
    for row in list(ws.iter_rows(values_only=True))[1:]:
        d = dict(zip(COL_HEADERS, row, strict=True))
        assert d["AI 리스크"] in (None, "")
        assert d["AI 이유"] in (None, "")


def test_export_mode_both_fills_both(seeded_client: TestClient) -> None:
    r = seeded_client.get("/documents/doc-1/export?mode=both")
    assert r.status_code == 200
    # r1은 AI 추천, r2는 사람 판정 — 동시에 보이는지
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    saw_ai = saw_human = False
    for name in wb.sheetnames:
        if name == "총괄표":
            continue
        ws = wb[name]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            d = dict(zip(COL_HEADERS, row, strict=True))
            if d["AI 리스크"]:
                saw_ai = True
            if d["사람 판정"]:
                saw_human = True
    assert saw_ai and saw_human


def test_export_404_for_unknown_doc(seeded_client: TestClient) -> None:
    r = seeded_client.get("/documents/no-such/export")
    assert r.status_code == 404


def test_export_409_when_no_requirements(seeded_client: TestClient) -> None:
    # 빈 doc-2 — Document만 등록하고 requirements 없음
    import asyncio

    async def seed_empty() -> None:
        await seeded_client.app.state.container.repo.save_document(
            Document(id="doc-2", src_path=Path("./_empty.pdf"), mime=DocumentMime.PDF)
        )

    asyncio.run(seed_empty())
    r = seeded_client.get("/documents/doc-2/export")
    assert r.status_code == 409
