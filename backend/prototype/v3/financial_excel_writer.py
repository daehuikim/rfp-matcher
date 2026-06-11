"""
금융 RFP 조견표 — 요건구분 + 카테고리 + 상세내용 + 출처.
"""
from __future__ import annotations

from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from prototype.v2.excel_writer import (
    _BODY_FONT,
    _BORDER,
    _HEADER_FILL,
    _HEADER_FONT,
    _CENTER,
    _WRAP,
    _WRAP_TOP,
    _ZEBRA_FILL,
    _merge_runs,
    _row_height,
    _safe_sheet,
    _write_overview,
)
from prototype.v2.extract import Req

FIN_COLUMNS = [
    ("요건구분", "group", 28, False, True),
    ("상세내용", "detail", 72, False, False),
    ("카테고리", "category", 44, False, True),
    ("출처", "source", 36, False, True),
]


def _write_financial_sheet(ws, reqs: list[Req]) -> None:
    for ci, (label, _k, width, _c, _m) in enumerate(FIN_COLUMNS, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    for ri, r in enumerate(reqs, 2):
        vals = {
            "group": (r.top or "").strip(),
            "category": (r.section_path or "").strip(),
            "source": (r.source or "").strip(),
            "detail": r.detail,
        }
        for ci, (_l, key, _w, center, _m) in enumerate(FIN_COLUMNS, 1):
            cell = ws.cell(ri, ci, vals[key])
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _CENTER if center else _WRAP
            if ri % 2 == 1:
                cell.fill = _ZEBRA_FILL
        h = _row_height(r.detail)
        if h:
            ws.row_dimensions[ri].height = h

    _merge_runs(ws, 1, lambda r: (r.top, r.section_path), reqs)
    _merge_runs(ws, 3, lambda r: (r.top, r.section_path), reqs)
    ws.freeze_panes = "A2"
    if reqs:
        ws.auto_filter.ref = f"A1:D{len(reqs) + 1}"


def write_financial_excel(
    reqs: list[Req],
    path,
    overview: dict | None = None,
    tab_order: list[str] | None = None,
) -> None:
    by_tab: OrderedDict[str, list[Req]] = OrderedDict()
    for r in reqs:
        by_tab.setdefault(r.tab or "요구사항", []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    if overview:
        ov_ws = wb.create_sheet(title=_safe_sheet("개요", used))
        _write_overview(ov_ws, overview)

    ordered = list(by_tab.keys())
    if tab_order:
        rank = {t: i for i, t in enumerate(tab_order)}
        ordered.sort(key=lambda t: (rank.get(t, 999), t))

    for tab in ordered:
        ws = wb.create_sheet(title=_safe_sheet(tab, used))
        _write_financial_sheet(ws, by_tab[tab])
    if not by_tab and not overview:
        wb.create_sheet(title="요구사항")
    wb.save(path)
