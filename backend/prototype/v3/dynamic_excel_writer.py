"""계위 깊이에 맞춘 동적 칼럼 조견표 writer — 고정 칼럼 없음.

각 Req.levels(외→내 계위 값)로 시트별 깊이를 정해
[요구사항 ID | <계위 칼럼들> | 상세요건 | 출처] 를 만든다.
levels 비면 [top, mid]에서 폴백. 계위 칼럼명은 level_names > 기본 사다리명.
스타일(헤더음영/테두리/줄바꿈/병합/줄높이/필터)은 v2 excel_writer 재사용.
"""
from __future__ import annotations

from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from prototype.v2.excel_writer import (
    _AI_FILL,
    _BODY_FONT,
    _BORDER,
    _CENTER,
    _HEADER_FILL,
    _HEADER_FONT,
    _WRAP,
    _WRAP_TOP,
    _ZEBRA_FILL,
    append_ai_columns,
    _row_height,
    _safe_sheet,
    _write_overview,
)
from prototype.v2.extract import Req


def _levels_of(r: Req) -> list[str]:
    """Req의 계위 값(외→내) — **위치 보존**. top="" + mid="X" → ["","X"](요구사항칸).

    중간 빈칸은 병합셀(상위 상속)이라 위치를 유지하고, 트레일링 빈칸만 제거해 깊이 산정.
    """
    if r.levels:
        lv = [(x or "").strip() for x in r.levels]
    else:
        lv = [(r.top or "").strip(), (r.mid or "").strip()]
    while lv and not lv[-1]:
        lv.pop()
    return lv


def _level_headers(depth: int, names: list[str]) -> list[str]:
    """계위 칼럼명 — 추출이 준 level_names 우선, 없으면 깊이기반 사다리.

    안쪽 2개 = 항목명/요구사항(라벨러 계위), 바깥 = 대분류/중분류…(섹션 계위).
    """
    if names and len(names) >= depth:
        return list(names[:depth])
    if depth <= 0:
        return []
    if depth == 1:
        return ["항목명"]
    if depth == 2:
        return ["항목명", "요구사항"]
    outer = depth - 2
    outer_names = ["대분류"] + [f"중분류{i}" for i in range(1, outer)]
    return [*outer_names, "항목명", "요구사항"]


def _clean_src(s: str) -> str:
    """출처 중복 세그먼트 정리: 'p.6 · 리스트 · 리스트' → 'p.6 · 리스트'."""
    parts = [p.strip() for p in (s or "").split("·")]
    out: list[str] = []
    for p in parts:
        if p and (not out or out[-1] != p):
            out.append(p)
    return " · ".join(out)


def _tab_depth(reqs: list[Req]) -> tuple[int, list[str]]:
    depth = 0
    names: list[str] = []
    for r in reqs:
        n = len(_levels_of(r))
        if n > depth:
            depth = n
        if r.level_names and len(r.level_names) > len(names):
            names = list(r.level_names)
    return depth, names


def _write_sheet(ws, reqs: list[Req], ai_by_rid: dict | None = None) -> None:
    depth, names = _tab_depth(reqs)
    headers = _level_headers(depth, names)
    columns = ["요구사항 ID", *headers, "상세요건", "출처"]
    widths = [16, *([24] * depth), 82, 16]
    for ci, label in enumerate(columns, 1):
        c = ws.cell(1, ci, label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = widths[ci - 1]
    ws.row_dimensions[1].height = 24

    lvl_col0 = 2
    src_col = 3 + depth

    # 계위 forward-fill: continuation("")은 상위 상속, 상위 값이 바뀌면 하위 reset.
    # → 자연스러운 셀 병합 준비(빈칸 강제 X, 이어지는 값만 병합).
    filled: list[list[str]] = []
    carry = [""] * depth
    for r in reqs:
        lv = (_levels_of(r) + [""] * depth)[:depth]
        row: list[str] = []
        for k in range(depth):
            if lv[k]:
                carry[k] = lv[k]
                for j in range(k + 1, depth):
                    carry[j] = ""
            row.append(carry[k])
        filled.append(row)

    for ri, (r, lv) in enumerate(zip(reqs, filled), 2):
        zebra = ri % 2 == 1
        row_vals = [r.rid, *lv, r.detail, _clean_src(r.source)]
        top_v = (r.top or "").strip()
        mid_v = (r.mid or "").strip()
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _CENTER if (ci == 1 or ci == src_col) else _WRAP
            # LLM 생성 셀(주황): ID(gen_rid) + 계위셀 중 본인의 top/mid 값과 일치하는 칸.
            # forward-fill 상속칸은 본인 top/mid와 불일치 → 일반색(앵커칸만 주황, 병합 안전).
            gen = (
                (ci == 1 and r.gen_rid)
                or (2 <= ci <= 1 + depth and val and r.gen_mid and val == mid_v)
                or (2 <= ci <= 1 + depth and val and r.gen_top and val == top_v)
            )
            if gen:
                cell.fill = _AI_FILL
            elif zebra:
                cell.fill = _ZEBRA_FILL
        h = _row_height(r.detail)
        if h:
            ws.row_dimensions[ri].height = h

    # prefix(외→k)가 같은 연속 행을 병합 — 값이 이어질 때만(자연 병합)
    n = len(filled)
    for k in range(depth):
        start = 0
        while start < n:
            if not filled[start][k]:
                start += 1
                continue
            end = start
            while end + 1 < n and filled[end + 1][: k + 1] == filled[start][: k + 1]:
                end += 1
            if end > start:
                ws.merge_cells(
                    start_row=start + 2, end_row=end + 2,
                    start_column=lvl_col0 + k, end_column=lvl_col0 + k,
                )
                ws.cell(start + 2, lvl_col0 + k).alignment = _WRAP_TOP
            start = end + 1

    ws.freeze_panes = "A2"
    if reqs:
        ws.auto_filter.ref = f"A1:{get_column_letter(src_col)}{len(reqs) + 1}"
    # 조견표 오른쪽에 AI 추천·사람 판정 칼럼(KT보유기술/부족기술/AI판정/AI설명/컨소시엄/Human판정/메모)
    append_ai_columns(ws, reqs, ai_by_rid, src_col)


def write_dynamic_excel(
    reqs: list[Req],
    path,
    overview: dict | None = None,
    tab_order: list[str] | None = None,  # 미사용 — 페이지순 보존을 위해 contiguous run 으로 시트 생성
    ai_by_rid: dict | None = None,
) -> None:
    # reqs 는 **페이지순**으로 들어온다. 같은 tab 의 page 연속 run 단위로 시트를 만든다 —
    # 비인접 동일 도메인은 별도 시트로(이름 _2 자동 dedup) 두어 **페이지순을 절대 흐트러뜨리지 않는다**.
    # (도메인별로 전역 묶으면 페이지가 뒤섞임 → 사람이 페이지순으로 못 읽음.)
    runs: list[tuple[str, list[Req]]] = []
    for r in reqs:
        t = r.tab or "요구사항"
        if runs and runs[-1][0] == t:
            runs[-1][1].append(r)
        else:
            runs.append((t, [r]))

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    if overview:
        ov = wb.create_sheet(title=_safe_sheet("개요", used))
        _write_overview(ov, overview)

    for tab, group in runs:
        ws = wb.create_sheet(title=_safe_sheet(tab, used))
        _write_sheet(ws, group, ai_by_rid)
    if not runs and not overview:
        wb.create_sheet(title="요구사항")
    wb.save(path)
