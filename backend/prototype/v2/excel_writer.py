"""
스타일 적용 엑셀 출력 — 정답 조견표 구조 재현.

탭(시트)별 분리 + 3계위 컬럼 [ID | 항목명 | 요구사항 | 상세요건 | 출처].
항목명·요구사항은 정답처럼 연속 동일값을 셀 병합. 헤더 음영·틀고정,
데이터 테두리·줄바꿈·세로가운데, 자동 행높이, 자동 필터.
"""
from __future__ import annotations

import math
import re
from collections import OrderedDict
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .extract import Req

# (헤더, Req 속성, 너비, 가운데정렬, 병합여부)
COLUMNS = [
    ("요구사항 ID", "rid", 16, True, False),
    ("항목명", "top", 22, False, True),
    ("요구사항", "mid", 26, False, True),
    ("상세요건", "detail", 82, False, False),
    ("산출정보", "deliverable", 22, False, True),
    ("관련요구사항", "related_req", 18, True, True),
    ("출처", "source", 16, True, False),
]
_DETAIL_COL = next(i for i, (_l, k, _w, _c, _m) in enumerate(COLUMNS, 1) if k == "detail")

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ZEBRA_FILL = PatternFill("solid", fgColor="EEF3FB")
_AI_FILL = PatternFill("solid", fgColor="FCE4D6")   # AI 생성 셀(ID 등) 구분색(주황)
_BODY_FONT = Font(size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="center")
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DETAIL_W = next(w for _l, k, w, _c, _m in COLUMNS if k == "detail")
_KO_PER_LINE = max(10, int(_DETAIL_W / 2.0))
_INVALID_SHEET = re.compile(r"[\\/?*\[\]:]")


def _safe_sheet(name: str, used: set[str]) -> str:
    base = (_INVALID_SHEET.sub("_", name).strip() or "요구사항")[:28]
    name, i = base, 2
    while name in used:
        name = f"{base}_{i}"[:31]
        i += 1
    used.add(name)
    return name


def _row_height(detail: str) -> float | None:
    if not detail:
        return None
    lines = sum(max(1, math.ceil(len(seg) / _KO_PER_LINE)) for seg in detail.split("\n"))
    return min(15 + (lines - 1) * 14, 260) if lines > 1 else None


def _merge_runs(ws, col_idx: int, key_fn, reqs: list[Req]) -> None:
    """연속 동일 key 구간을 셀 병합."""
    start = 0
    n = len(reqs)
    while start < n:
        end = start
        while end + 1 < n and key_fn(reqs[end + 1]) == key_fn(reqs[start]) and key_fn(reqs[start]):
            end += 1
        if end > start:
            ws.merge_cells(start_row=start + 2, end_row=end + 2,
                           start_column=col_idx, end_column=col_idx)
            ws.cell(start + 2, col_idx).alignment = _WRAP_TOP
        start = end + 1


def _expand_image_rows(reqs: list[Req]) -> list[Req]:
    """텍스트와 이미지를 다른 행으로 — 셀 겹침 방지."""
    out: list[Req] = []
    for r in reqs:
        imgs = r.detail_images
        if not imgs:
            out.append(r)
            continue
        detail = (r.detail or "").strip()
        is_caption = detail.startswith("[관련 화면(안)]") or detail == "[표]"
        if detail and not is_caption:
            t = copy(r)
            t.detail_images = []
            out.append(t)
            first_detail = "[표]"
        elif is_caption:
            first_detail = detail
        else:
            first_detail = "[관련 화면(안)]"
        for i, path in enumerate(imgs):
            ir = copy(r)
            ir.detail = first_detail if i == 0 else ""
            ir.detail_images = [path]
            out.append(ir)
            first_detail = ""
    return out


def _embed_detail_images(ws, ri: int, paths: list[str], *, text_lines: int = 0) -> None:
    """상세요건 셀 — 이미지 전용 행에 삽입(텍스트와 겹치지 않음)."""
    if not paths:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return
    col = get_column_letter(_DETAIL_COL)
    row_h = ws.row_dimensions[ri].height or 15
    y_off = 0
    if text_lines:
        y_off = min(14 * text_lines, 120)
    for path in paths:
        try:
            img = XLImage(path)
        except Exception:
            continue
        scale = min(1.0, 520 / max(img.width, 1), 480 / max(img.height, 1))
        img.width = max(120, int(img.width * scale))
        img.height = max(80, int(img.height * scale))
        img.anchor = f"{col}{ri}"
        ws.add_image(img)
        need = img.height * 0.75 + 8 + y_off
        row_h = max(row_h, need)
        y_off = 0
    ws.row_dimensions[ri].height = row_h


# ── AI 판정·사람 판정 칼럼 (조견표 오른쪽에 덧붙임) ──────────────────────────
_AI_HEADERS = ["KT 보유 기술", "부족 기술", "AI 판정", "AI 설명", "컨소시엄 필요 사항", "Human 판정", "Human 메모"]
_AI_WIDTHS = [26, 22, 9, 52, 24, 11, 28]
_AI_CENTER = {2, 5}  # AI 판정, Human 판정 → 가운데 정렬
_HUMAN_START = 5  # 이 인덱스부터 사람(Human) 칼럼 — 고유색
_AI_HDR_FILL = PatternFill("solid", fgColor="ED7D31")     # AI 칼럼 헤더(주황 — AI가 한 것)
_HUMAN_HDR_FILL = PatternFill("solid", fgColor="548235")  # 사람 칼럼 헤더(초록 — 사람이 입력)
_HUMAN_BODY_FILL = PatternFill("solid", fgColor="E2EFDA")  # 사람 칼럼 본문(연초록)


def _ai_row_values(rec, jud) -> list[str]:
    """ai_by_rid 값 (rec, jud) → 7개 셀 문자열. rec/jud는 duck-typed."""

    def _enum(v) -> str:
        return str(getattr(v, "value", v) or "")

    if rec is None:
        kt = miss = risk = reason = consortium = ""
    else:
        kt = (getattr(rec, "related_solution", "") or "").strip()
        if not kt:
            kt = ", ".join(getattr(rec, "matched_solutions", []) or [])
        miss = ", ".join(getattr(rec, "missing_tech", []) or [])
        risk = _enum(getattr(rec, "ai_risk", ""))
        reason = getattr(rec, "ai_reason", "") or ""
        consortium = getattr(rec, "consortium_need", "") or ""
    mark = _enum(getattr(jud, "mark", "")) if jud is not None else ""
    note = (getattr(jud, "note", "") or "") if jud is not None else ""
    return [kt, miss, risk, reason, consortium, mark, note]


def append_ai_columns(ws, reqs, ai_by_rid, base_cols: int) -> None:
    """조견표 시트 오른쪽에 AI/사람 판정 칼럼을 덧붙인다. ai_by_rid={rid:(rec,jud)}."""
    if not ai_by_rid:
        return
    for j, label in enumerate(_AI_HEADERS):
        ci = base_cols + 1 + j
        cell = ws.cell(1, ci, label)
        cell.fill = _HUMAN_HDR_FILL if j >= _HUMAN_START else _AI_HDR_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = _AI_WIDTHS[j]
    for ri, r in enumerate(reqs, 2):
        rec, jud = ai_by_rid.get(id(r), (None, None))  # 객체 id로 행 단위 매핑 (중복 rid 안전)
        for j, val in enumerate(_ai_row_values(rec, jud)):
            ci = base_cols + 1 + j
            cell = ws.cell(ri, ci, val)
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _CENTER if j in _AI_CENTER else _WRAP
            # 칼럼 고유색: AI(주황) vs 사람(초록)으로 출처 구분
            cell.fill = _HUMAN_BODY_FILL if j >= _HUMAN_START else _AI_FILL
    if reqs:
        ws.auto_filter.ref = f"A1:{get_column_letter(base_cols + len(_AI_HEADERS))}{len(reqs) + 1}"


def _write_sheet(ws, reqs: list[Req], ai_by_rid: dict | None = None) -> None:
    reqs = _expand_image_rows(reqs)
    for ci, (label, _k, width, _c, _m) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    for ri, r in enumerate(reqs, 2):
        vals = {
            "rid": r.rid,
            "top": r.top,
            "mid": r.mid,
            "detail": r.detail,
            "deliverable": r.deliverable,
            "related_req": r.related_req,
            "source": r.source,
        }
        gen = {"rid": r.gen_rid, "top": r.gen_top, "mid": r.gen_mid}
        for ci, (_l, key, _w, center, _m) in enumerate(COLUMNS, 1):
            cell = ws.cell(ri, ci, vals[key])
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _CENTER if center else _WRAP
            if gen.get(key):          # AI 가 생성한 값만 주황 (원문 추출은 일반색)
                cell.fill = _AI_FILL
            elif ri % 2 == 1:
                cell.fill = _ZEBRA_FILL
        h = _row_height(r.detail)
        if h:
            ws.row_dimensions[ri].height = h
        if r.detail_images:
            _embed_detail_images(ws, ri, r.detail_images)

    # ID(top+mid) / 항목명(top) / 요구사항(top+mid) 셀 병합
    _merge_runs(ws, 1, lambda r: r.rid, reqs)
    _merge_runs(ws, 2, lambda r: r.top, reqs)
    _merge_runs(ws, 3, lambda r: (r.top, r.mid), reqs)
    _merge_runs(ws, 5, lambda r: (r.rid, r.deliverable), reqs)
    _merge_runs(ws, 6, lambda r: (r.rid, r.related_req), reqs)

    ws.freeze_panes = "A2"
    if reqs:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(reqs) + 1}"
    append_ai_columns(ws, reqs, ai_by_rid, len(COLUMNS))


_TITLE_FONT = Font(bold=True, size=14, color="305496")
_SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
_SEC_FONT = Font(bold=True, size=11, color="1F3864")


def _write_public_rfp_overview(ws, overview: dict) -> None:
    """개요(좌) + 빈열 + 요구사항 총괄표(우) — 시트명 '개요'."""
    meta = overview.get("meta") or {}
    headers = overview.get("headers") or []
    rows = overview.get("rows") or []

    left_cols = 3
    gap_col = 4
    right_start = 5

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions[get_column_letter(gap_col)].width = 2
    for ci, label in enumerate(headers, right_start):
        ws.column_dimensions[get_column_letter(ci)].width = 24 if ci == right_start else 20

    ws.cell(1, 1, "RFP 개요").font = _TITLE_FONT
    ws.cell(1, right_start, "요구사항 총괄표").font = _TITLE_FONT

    row = 3

    def section(title: str, start_row: int) -> int:
        c = ws.cell(start_row, 1, title)
        c.fill = _SEC_FILL
        c.font = _SEC_FONT
        ws.cell(start_row, 2).fill = _SEC_FILL
        ws.cell(start_row, 3).fill = _SEC_FILL
        return start_row + 1

    row = section("전체 요약", row)
    summary = meta.get("summary", "")
    cell = ws.cell(row, 1, summary)
    cell.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    h = _row_height(summary)
    if h:
        ws.row_dimensions[row].height = max(h, 60)
    row += 2

    row = section("핵심 기술", row)
    for ci, lbl in enumerate(("기술", "요구", "관련 ID"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = Font(bold=True, size=10)
        c.fill = _SEC_FILL
    row += 1
    for name, req, ids in meta.get("techs", []):
        ws.cell(row, 1, name).font = Font(bold=True, size=10)
        ws.cell(row, 1).alignment = _WRAP
        ws.cell(row, 2, req).alignment = _WRAP
        ws.cell(row, 3, ids).fill = _AI_FILL
        ws.cell(row, 3).alignment = _WRAP
        for ci in (1, 2, 3):
            ws.cell(row, ci).border = _BORDER
        row += 1
    row += 1

    row = section("핵심 RISK (독소조항)", row)
    for ci, lbl in enumerate(("관련 ID", "리스크/독소조항"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = Font(bold=True, size=10)
        c.fill = _SEC_FILL
    ws.cell(row, 2).fill = _SEC_FILL
    row += 1
    risks = meta.get("risks", [])
    if not risks:
        ws.cell(row, 1, "(식별된 독소조항 없음)").font = Font(italic=True, size=10, color="808080")
        row += 1
    for clause, ids in risks:
        ws.cell(row, 1, ids).alignment = _WRAP
        ws.cell(row, 1).fill = _AI_FILL
        ws.cell(row, 2, clause).alignment = _WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row, 1).border = _BORDER
        ws.cell(row, 2).border = _BORDER
        row += 1

    left_end = max(row, 3)

    for ci, label in enumerate(headers, right_start):
        cell = ws.cell(3, ci, label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    for ri, data in enumerate(rows, 4):
        for ci, val in enumerate(data, right_start):
            cell = ws.cell(ri, ci, val)
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _WRAP
            if ri % 2 == 0:
                cell.fill = _ZEBRA_FILL
    right_end = 3 + len(rows)

    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = (
            f"{get_column_letter(right_start)}3:"
            f"{get_column_letter(right_start + max(len(headers), 1) - 1)}{right_end}"
        )
    # 좌·우 블록 높이 맞춤
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 8


def _write_overview(ws, overview: dict) -> None:
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 24
    ws.cell(1, 1, "RFP 개요").font = _TITLE_FONT
    row = 3

    def section(title: str):
        nonlocal row
        c = ws.cell(row, 1, title)
        c.fill = _SEC_FILL
        c.font = _SEC_FONT
        ws.cell(row, 2).fill = _SEC_FILL
        row += 1

    section("전체 요약")
    cell = ws.cell(row, 1, overview.get("summary", ""))
    cell.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    h = _row_height(overview.get("summary", ""))
    if h:
        ws.row_dimensions[row].height = max(h, 60)
    row += 2

    section("핵심 기술")
    for ci, lbl in enumerate(("기술", "요구", "관련 요구사항 ID"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = Font(bold=True, size=10)
        c.fill = _SEC_FILL
    row += 1
    for name, req, ids in overview.get("techs", []):
        ws.cell(row, 1, name).font = Font(bold=True, size=10)
        ws.cell(row, 1).alignment = _WRAP
        ws.cell(row, 2, req).alignment = _WRAP
        ws.cell(row, 3, ids).fill = _AI_FILL
        ws.cell(row, 3).alignment = _WRAP
        for ci in (1, 2, 3):
            ws.cell(row, ci).border = _BORDER
        row += 1
    row += 1

    section("핵심 RISK (독소조항)")
    for ci, lbl in enumerate(("관련 요구사항 ID", "리스크/독소조항"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = Font(bold=True, size=10)
        c.fill = _SEC_FILL
    row += 1
    risks = overview.get("risks", [])
    if not risks:
        ws.cell(row, 1, "(식별된 독소조항 없음)").font = Font(italic=True, size=10, color="808080")
        row += 1
    for clause, ids in risks:
        ws.cell(row, 1, ids).alignment = _WRAP      # ID 먼저
        ws.cell(row, 1).fill = _AI_FILL
        ws.cell(row, 2, clause).alignment = _WRAP    # 내용 나중
        ws.cell(row, 1).border = _BORDER
        ws.cell(row, 2).border = _BORDER
        row += 1


def write_excel(
    reqs: list[Req],
    path,
    overview: dict | None = None,
    tab_order: list[str] | None = None,
    ai_by_rid: dict | None = None,
) -> None:
    by_tab: "OrderedDict[str, list[Req]]" = OrderedDict()
    for r in reqs:
        by_tab.setdefault(r.tab or "요구사항", []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    if overview and overview.get("type") == "public_rfp":
        ov_ws = wb.create_sheet(title=_safe_sheet("개요", used))
        _write_public_rfp_overview(ov_ws, overview)
    elif overview and overview.get("type") == "summary_table":
        title = overview.get("sheet_title") or "요구사항 총괄표"
        ov_ws = wb.create_sheet(title=_safe_sheet(title, used))
        _write_summary_table(ov_ws, overview)
    elif overview:
        ov_ws = wb.create_sheet(title="개요")
        _write_overview(ov_ws, overview)
        used.add("개요")

    ordered_tabs = list(by_tab.keys())
    if tab_order:
        rank = {t: i for i, t in enumerate(tab_order)}
        ordered_tabs.sort(key=lambda t: (rank.get(t, 999 if t != "부록" else 1000), t))

    for tab in ordered_tabs:
        items = by_tab[tab]
        ws = wb.create_sheet(title=_safe_sheet(tab, used))
        _write_sheet(ws, items, ai_by_rid)
    if not by_tab and not overview:
        wb.create_sheet(title="요구사항")
    wb.save(path)


def _write_summary_table(ws, overview: dict) -> None:
    headers = overview.get("headers") or []
    rows = overview.get("rows") or []
    for ci, label in enumerate(headers, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 28 if ci == 1 else 22
    ws.row_dimensions[1].height = 24
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.border = _BORDER
            cell.font = _BODY_FONT
            cell.alignment = _WRAP
            if ri % 2 == 0:
                cell.fill = _ZEBRA_FILL
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(max(len(headers), 1))}{len(rows) + 1}"
