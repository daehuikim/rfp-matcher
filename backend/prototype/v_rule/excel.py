"""스펙 5단계 — 조견표 db → Excel. Section=탭, [요구사항 ID·항목명·요구사항·상세요건·출처].

앱 import(parse_excel) 호환 포맷 — 업로드하면 AI 매칭+예쁜 재export 가 그대로 동작.
참조사항은 출처 칼럼에 합쳐 보존(레벨 칼럼 오염 방지).
"""
from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .db import Row

_HDR = Font(bold=True, size=10, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor="404040")
_WRAP = Alignment(wrap_text=True, vertical="top")
_BORDER = Border(*[Side(style="thin", color="D0D0D0")] * 4)
_COLS = ["요구사항 ID", "항목명", "요구사항", "상세요건", "출처"]


def _safe(title: str, used: set[str]) -> str:
    t = re.sub(r"[\\/*?:\[\]]", " ", title or "요구사항").strip()[:31] or "요구사항"
    base, i = t, 2
    while t in used:
        t = f"{base[:28]}_{i}"; i += 1
    used.add(t)
    return t


def _slug(text: str) -> str:
    toks = re.findall(r"[0-9A-Za-z가-힣]+", text or "")
    return ("".join(toks)[:20]) or "요구사항"


def assign_rids(rows: list[Row]) -> None:
    """탭(Section) 기반 ID — 한 탭 = 한 접두사 + 연속번호 (앱 규칙과 동일 철학)."""
    counters: dict[str, int] = {}
    prefix: dict[str, str] = {}
    used: set[str] = set()
    for r in rows:
        sec = r.section or "요구사항"
        if sec not in prefix:
            base = _slug(sec); pfx, k = base, 2
            while pfx in used:
                pfx = f"{base}{k}"; k += 1
            used.add(pfx); prefix[sec] = pfx
        p = prefix[sec]
        counters[p] = counters.get(p, 0) + 1
        r._rid = f"{p}_{counters[p]:03d}"  # type: ignore[attr-defined]


def write_excel(rows: list[Row], out_path: str | Path) -> dict:
    """rows → Excel(앱 호환). 반환: 통계(시트수·행수)."""
    assign_rids(rows)
    by_section: "OrderedDict[str, list[Row]]" = OrderedDict()
    for r in rows:
        by_section.setdefault(r.section or "요구사항", []).append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for section, srows in by_section.items():
        ws = wb.create_sheet(_safe(section, used))
        for ci, h in enumerate(_COLS, 1):
            c = ws.cell(1, ci, h); c.font = _HDR; c.fill = _HDR_FILL
        for ci, w in enumerate([16, 22, 26, 70, 16], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        prev_item = prev_req = None
        for ri, r in enumerate(srows, 2):
            item = "" if r.item == prev_item and r.item else r.item
            req = "" if (r.requirement == prev_req and r.requirement and not item) else r.requirement
            src = "; ".join(x for x in [f"p.{r.page}" if r.page else "", f"참조:{r.ref}" if r.ref else ""] if x)
            vals = [getattr(r, "_rid", ""), item, req, r.detail, src]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(ri, ci, v); cell.alignment = _WRAP; cell.border = _BORDER
            prev_item, prev_req = r.item, r.requirement
    if not by_section:
        wb.create_sheet("요구사항")
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"sheets": len(by_section), "rows": len(rows)}
