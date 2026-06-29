"""조견표 Excel(.xlsx) → Requirement/Recommendation/HumanJudgement 역파싱.

export(write_dynamic_excel/write_excel) 포맷을 되읽는다:
  시트=탭, 헤더 [요구사항 ID | <계위..> | 상세요건 | 출처 | (KT 보유 기술 | 부족 기술 | AI 판정 |
  AI 설명 | 컨소시엄 필요 사항 | Human 판정 | Human 메모)].
- **편집본 재업로드**: AI/Human 판정 칼럼이 있으면 복원(라운드트립).
- **외부 조견표 불러오기**: 추출 칼럼만 있어도 OK(판정은 비고 AI 매칭으로 채울 수 있음).
계위 칼럼 수는 가변(동적 writer) — ID/상세요건/출처/AI칼럼을 제외한 가운데 칼럼들을 계위로 본다.
"""
from __future__ import annotations

import re
import uuid
from pathlib import Path

import openpyxl

from app.domain.enums import CategorySource, Judgement
from app.domain.models import HumanJudgement, Recommendation, Requirement
from prototype.v2.extract import Req as V2Req

_AI_HEADERS = {
    "KT 보유 기술", "부족 기술", "AI 판정", "AI 설명",
    "컨소시엄 필요 사항", "Human 판정", "Human 메모",
}
_FIXED = {"요구사항 ID", "상세요건", "출처"} | _AI_HEADERS
_JUDG = {"O": Judgement.YES, "△": Judgement.PARTIAL, "X": Judgement.NO, "": Judgement.UNSET}


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _judg(v) -> Judgement:
    return _JUDG.get(_s(v), Judgement.UNSET)


def _parse_overview(wb) -> dict | None:
    """'개요' 시트 → overview dict (재export 라운드트립). _write_overview 역연산.

    레이아웃: 'RFP 개요' / '전체 요약'→summary / '핵심 기술'→[기술,요구,ID] 행 /
    '핵심 RISK (독소조항)'→[ID, 리스크/독소조항] 행(writer 가 ID먼저·내용나중으로 씀).
    """
    if "개요" not in wb.sheetnames:
        return None
    ws = wb["개요"]
    summary = ""
    techs: list[tuple[str, str, str]] = []
    risks: list[tuple[str, str]] = []
    section: str | None = None
    for row in ws.iter_rows(values_only=True):
        a = _s(row[0]) if len(row) > 0 else ""
        b = _s(row[1]) if len(row) > 1 else ""
        c = _s(row[2]) if len(row) > 2 else ""
        if a == "전체 요약":
            section = "summary"; continue
        if a == "핵심 기술":
            section = "techs"; continue
        if a.startswith("핵심 RISK"):
            section = "risks"; continue
        if not (a or b or c):
            continue
        if section == "summary" and a and a != "RFP 개요":
            summary = a
        elif section == "techs":
            if a == "기술" and b == "요구":   # 헤더행 스킵
                continue
            if a or b:
                techs.append((a, b, c))       # (기술, 요구, 관련ID)
        elif section == "risks":
            if a == "관련 요구사항 ID" or a == "(식별된 독소조항 없음)":
                continue
            if a or b:
                risks.append((b, a))          # 시트 [ID, 내용] → overview (clause, ids)
    if not (summary or techs or risks):
        return None
    return {"summary": summary, "techs": techs, "risks": risks}


def parse_excel(path: str | Path, doc_id: str) -> tuple[
    list[Requirement], dict[str, Recommendation], dict[str, HumanJudgement], list[V2Req], dict | None
]:
    """반환: (앱 Requirement, 추천, 사람판정, **v2 Req 리스트**, **overview dict**).

    v2 Req 는 reqs 와 1:1 동일 순서 — import 후 v3_export.pkl 로 저장하면 재export 가
    write_dynamic_excel(동적칼럼·AI칼럼)을 타서 **라운드트립 무손실**(탭순서·계위·AI칼럼 보존).
    overview 는 '개요' 시트를 역파싱 — 재export 시 개요 시트까지 복원(완전 라운드트립).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    overview = _parse_overview(wb)
    reqs: list[Requirement] = []
    recs: dict[str, Recommendation] = {}
    juds: dict[str, HumanJudgement] = {}
    v2reqs: list[V2Req] = []

    for name in wb.sheetnames:
        if name in ("개요", "요구사항 총괄표"):
            continue
        ws = wb[name]
        if ws.max_row < 2:
            continue
        hdr = [_s(c.value) for c in ws[1]]
        idx = {h: i for i, h in enumerate(hdr) if h}
        if "상세요건" not in idx:  # 조견표 시트가 아님(개요/서식 등) → 스킵
            continue
        level_cols = [i for i, h in enumerate(hdr) if h and h not in _FIXED]
        level_names = [hdr[i] for i in level_cols]  # 계위 칼럼 헤더(대분류/항목명/요구사항…) 재현용
        det_i, src_i, id_i = idx.get("상세요건"), idx.get("출처"), idx.get("요구사항 ID")
        has_ai = bool(_AI_HEADERS & set(idx))

        # 병합셀 forward-fill(상위 계위는 병합되어 빈칸) — 위에서부터 carry
        carry = [""] * len(hdr)
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(r, c + 1).value for c in range(len(hdr))]
            detail = _s(cells[det_i]) if det_i is not None else ""
            if not detail:
                continue
            for i in level_cols:  # 계위 병합 carry
                v = _s(cells[i])
                if v:
                    carry[i] = v
                    for j in level_cols:
                        if j > i:
                            carry[j] = ""
            levels = [carry[i] for i in level_cols]
            nonempty = [v for v in levels if v]
            name_v = nonempty[-2] if len(nonempty) >= 2 else (nonempty[-1] if nonempty else "")
            mid_v = nonempty[-1] if len(nonempty) >= 2 else ""
            rid = _s(cells[id_i]) if id_i is not None else ""

            rid = rid or f"{name[:4]}-{len(reqs) + 1:03d}"
            source = _s(cells[src_i]) if src_i is not None else ""
            req = Requirement(
                id=uuid.uuid4().hex,
                doc_id=doc_id,
                category=name,
                code=rid,
                name=name_v or detail[:30],
                definition=mid_v or None,
                detail=detail,
                source_ref=source or None,
                category_source=CategorySource.DOCUMENT_TABLE,
            )
            reqs.append(req)

            # 재export 무손실용 v2 Req(levels/level_names/tab/source 보존) — reqs 와 동일 순서
            pm = re.search(r"p\.(\d+)", source)
            v2reqs.append(V2Req(
                doc=name, table_id=(int(pm.group(1)) if pm else -1),
                page=(int(pm.group(1)) if pm else None),
                rid=rid, top=name_v, mid=mid_v, detail=detail, tab=name,
                source=source, levels=list(levels), level_names=list(level_names),
            ))

            if has_ai:
                kt = _s(cells[idx["KT 보유 기술"]]) if "KT 보유 기술" in idx else ""
                miss = _s(cells[idx["부족 기술"]]) if "부족 기술" in idx else ""
                reason = _s(cells[idx["AI 설명"]]) if "AI 설명" in idx else ""
                cons = _s(cells[idx["컨소시엄 필요 사항"]]) if "컨소시엄 필요 사항" in idx else ""
                risk = _judg(cells[idx["AI 판정"]]) if "AI 판정" in idx else Judgement.UNSET
                if risk != Judgement.UNSET or kt or reason:
                    recs[req.id] = Recommendation(
                        requirement_id=req.id, ai_risk=risk, ai_reason=reason,
                        related_solution=kt,
                        missing_tech=[m.strip() for m in miss.split(",") if m.strip()],
                        consortium_need=cons or None,
                    )
                hmark = _judg(cells[idx["Human 판정"]]) if "Human 판정" in idx else Judgement.UNSET
                hnote = _s(cells[idx["Human 메모"]]) if "Human 메모" in idx else ""
                if hmark != Judgement.UNSET or hnote:
                    juds[req.id] = HumanJudgement(
                        requirement_id=req.id, mark=hmark, note=hnote,
                    )
    return reqs, recs, juds, v2reqs, overview
