from __future__ import annotations

import re
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import Workbook

from app.domain.enums import ExportMode, Judgement
from app.domain.models import ExtractionMetadata, HumanJudgement, Recommendation, Requirement
from app.phase1.extraction.category_provenance import (
    category_source_label,
    document_category_spec,
    summarize_category_sources,
)

from .export_columns import EXPORT_COLUMNS, column_headers, resolve_export_columns
from .sheet_style import merge_hierarchy_columns, style_data_sheet, write_overview_sheet

_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def _safe_sheet_name(name: str) -> str:
    name = _INVALID_SHEET_CHARS.sub("_", name).strip() or "기타"
    return name[:31]


def _document_order_key(r: Requirement) -> tuple[int, int, str]:
    """RFP 원문 순서 정렬 키 — 페이지 → 표 인덱스 → 코드."""
    page = r.source_page
    try:
        page_i = int(page) if page is not None else 10**6
    except (TypeError, ValueError):
        page_i = 10**6
    table_i = r.source_table_index if isinstance(r.source_table_index, int) else 10**6
    return (page_i, table_i, r.code or "")


_TAB_NUM = re.compile(
    r"^\s*(?:[IVXLCDM]+[.)]|[가나다라마바사아자차카타파하][.)]|[Ⅰ-Ⅻ]\.?|\d+(?:\.\d+)*\.?)\s+"
)
_TAB_BULLET = re.compile(r"^\s*[❍○●▪∙•◦·\-–—]\s*")
_TAB_BRACKET = re.compile(r"\[[^\]]*\]|\([^)]*\)|「[^」]*」")


def _clean_tab_name(section_path: str) -> str:
    """문서 섹션 경로 → 엑셀 탭 이름(번호·괄호·목차 점선 제거)."""
    seg = section_path.split(" > ")[-1]
    seg = _TAB_BRACKET.sub("", seg)
    seg = re.sub(r"[·.·…‧\.\-_\s]{3,}\d*\s*$", "", seg)
    seg = re.sub(r"\s+\d+\s*$", "", seg)
    for _ in range(3):
        new = _TAB_BULLET.sub("", _TAB_NUM.sub("", seg))
        if new == seg:
            break
        seg = new
    seg = seg.strip()
    return seg[:40] or (section_path.split(" > ")[-1][:40] or "요구사항")


def _tab_sheet_key(r: Requirement) -> str:
    """시트(탭) 이름 — V2 LLM 도메인 탭(category)만 사용 (표#N 폴백 금지)."""
    tab = (r.category or "").strip()
    if tab and tab not in ("미분류", "기타"):
        return _clean_tab_name(tab)
    return "요구사항"


def _unique_sheet_name(base: str, used: set[str]) -> str:
    name = _safe_sheet_name(base)
    if name not in used:
        used.add(name)
        return name
    stem = name[:28]
    i = 2
    while True:
        candidate = f"{stem}_{i}"[:31]
        if candidate not in used:
            used.add(candidate)
            return candidate
        i += 1


def _group_by_tab_sheets(requirements: list[Requirement]) -> OrderedDict[str, list[Requirement]]:
    """도메인 탭(category)별 시트 — 신한 processed·results_final 과 동일 원칙."""
    groups: OrderedDict[str, list[Requirement]] = OrderedDict()
    used_names: set[str] = set()
    key_to_sheet: dict[str, str] = {}
    for r in sorted(requirements, key=_document_order_key):
        tab_key = _tab_sheet_key(r)
        if tab_key not in key_to_sheet:
            key_to_sheet[tab_key] = _unique_sheet_name(tab_key, used_names)
        groups.setdefault(key_to_sheet[tab_key], []).append(r)
    return groups


class RequirementSheetWriter:
    """
    Requirement + Recommendation + HumanJudgement → xlsx.

    cols 파라미터로 출력 칼럼을 선택할 수 있다 (code 등 시스템 생성 필드 제외 가능).
    """

    def write(
        self,
        out_path: Path,
        requirements: list[Requirement],
        recommendations: dict[str, Recommendation] | None = None,
        judgements: dict[str, HumanJudgement] | None = None,
        mode: ExportMode = ExportMode.BOTH,
        columns: list[str] | None = None,
        adaptive: bool = True,
        extraction_meta: ExtractionMetadata | None = None,
        layout: str = "cluster",
        v2_overview: dict | None = None,
    ) -> Path:
        """layout: "cluster"=분류별 시트(기술 중심) / "ordered"=RFP 원문 순서·섹션별 시트."""
        out_path.parent.mkdir(parents=True, exist_ok=True)
        recs = recommendations or {}
        judges = judgements or {}
        col_keys = resolve_export_columns(
            requirements,
            recs,
            judges,
            mode,
            columns,
            adaptive=adaptive,
        )
        headers = column_headers(col_keys)

        wb = Workbook()
        wb.remove(wb.active)

        by_cat: dict[str, list[Requirement]] = defaultdict(list)
        cat_sources: dict[str, str] = {}
        for r in requirements:
            cat = r.category or "기타"
            by_cat[cat].append(r)
            if cat not in cat_sources:
                cat_sources[cat] = category_source_label(r.category_source)

        overview_title = (
            v2_overview.get("sheet_title") or "요구사항 총괄표"
            if v2_overview and v2_overview.get("type") == "summary_table"
            else "개요"
        )
        overview = wb.create_sheet(title=_safe_sheet_name(overview_title))
        write_overview_sheet(
            overview,
            requirements,
            recs,
            by_cat,
            document_category_spec(extraction_meta, requirements),
            v2_overview=v2_overview,
        )

        recs = recommendations or {}
        judges = judgements or {}

        if layout == "ordered":
            # RFP 원문 순서 — 도메인 탭별 시트(신한·processed 형식), 시트 내 원문 순 정렬
            for tab_name, items in _group_by_tab_sheets(requirements).items():
                ws = wb.create_sheet(title=tab_name)
                ws.append(headers)
                for r in items:
                    ws.append(self._build_row(r, recs.get(r.id), judges.get(r.id), col_keys, mode))
                style_data_sheet(
                    ws, col_keys, len(items),
                    gen_by_row=[r.ai_generated_fields for r in items],
                )
                merge_hierarchy_columns(
                    ws,
                    col_keys,
                    len(items),
                    names=[r.name for r in items],
                    definitions=[r.definition or "" for r in items],
                )
        else:
            # 도메인 탭(category)별 시트 — 시트 내 항목명·요구사항(중분류) 셀 병합
            for cat, items in sorted(by_cat.items()):
                ws = wb.create_sheet(title=_safe_sheet_name(cat))
                ws.append(headers)
                for r in items:
                    ws.append(self._build_row(r, recs.get(r.id), judges.get(r.id), col_keys, mode))
                style_data_sheet(
                    ws, col_keys, len(items),
                    gen_by_row=[r.ai_generated_fields for r in items],
                )
                merge_hierarchy_columns(
                    ws,
                    col_keys,
                    len(items),
                    names=[r.name for r in items],
                    definitions=[r.definition or "" for r in items],
                )

        wb.save(out_path)
        return out_path

    @staticmethod
    def _build_row(
        req: Requirement,
        rec: Recommendation | None,
        jud: HumanJudgement | None,
        col_keys: list[str],
        mode: ExportMode,
    ) -> list[str]:
        ai_keys = {
            "ai_risk",
            "ai_reason",
            "matched_solutions",
            "missing_tech",
            "consortium",
        }
        human_keys = {"human_mark", "human_note"}
        out: list[str] = []
        for key in col_keys:
            if key in ai_keys and mode == ExportMode.HUMAN:
                out.append("")
                continue
            if key in human_keys and mode == ExportMode.AI:
                out.append("" if key != "human_mark" else Judgement.UNSET.value)
                continue
            out.append(EXPORT_COLUMNS[key].cell_value(req, rec, jud))
        return out
