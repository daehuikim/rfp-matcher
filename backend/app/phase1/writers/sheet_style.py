"""
엑셀 출력 스타일 (단일 선언 소스).

조견표가 정답 엑셀처럼 보이도록 헤더 음영·볼드·틀고정, 데이터 테두리·줄바꿈·
세로 가운데, 컬럼 너비, 지브라 음영, 자동 행높이, 자동 필터를 적용한다.
컬럼 너비/색을 바꾸려면 이 파일의 선언부만 고친다.
"""
from __future__ import annotations

import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ZEBRA_FILL = PatternFill("solid", fgColor="EEF3FB")
AI_FILL = PatternFill("solid", fgColor="FCE4D6")  # AI 생성 셀(원문 추출 아님) 구분색(주황)
BODY_FONT = Font(size=10)
_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# export 컬럼 key -> (너비, 가운데정렬 여부)
COLUMN_STYLE: dict[str, tuple[int, bool]] = {
    "category": (18, False),
    "subcategory": (16, False),
    "code": (12, True),
    "name": (24, False),
    "definition": (50, False),
    "detail": (80, False),
    "source_page": (9, True),
    "source_section": (20, False),
    "deliverables": (40, False),
    "related": (18, False),
    "ai_risk": (10, True),
    "ai_reason": (50, False),
    "matched_solutions": (30, False),
    "missing_tech": (24, False),
    "consortium": (20, False),
    "human_mark": (10, True),
    "human_note": (30, False),
    "category_source": (12, True),
}
_DEFAULT = (18, False)
# 행높이를 좌우하는 긴 본문 컬럼
_TALL_KEYS = {"definition", "detail", "deliverables", "ai_reason", "matched_solutions", "human_note"}


def _row_height(ws: Worksheet, ri: int, col_keys: list[str]) -> float | None:
    lines = 1
    for ci, key in enumerate(col_keys, 1):
        if key not in _TALL_KEYS:
            continue
        val = ws.cell(ri, ci).value
        if not val:
            continue
        width, _ = COLUMN_STYLE.get(key, _DEFAULT)
        per_line = max(8, int(width / 2.0))  # 한글 ≈ 2배 폭
        n = 0
        for seg in str(val).split("\n"):
            n += max(1, math.ceil(len(seg) / per_line))
        lines = max(lines, n)
    if lines <= 1:
        return None
    return min(15 + (lines - 1) * 14, 260)


def style_data_sheet(
    ws: Worksheet,
    col_keys: list[str],
    n_rows: int,
    gen_by_row: list[set[str]] | None = None,
) -> None:
    """헤더 1행 + 데이터 n_rows 가 이미 append 된 시트에 스타일 적용.

    gen_by_row[i] 에 든 칼럼 key 셀은 AI 생성값이므로 주황(AI_FILL)으로 칠한다
    (지브라 음영보다 우선). 원문 추출 셀은 기존 지브라 규칙 유지.
    """
    for ci, key in enumerate(col_keys, 1):
        cell = ws.cell(1, ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        width, _ = COLUMN_STYLE.get(key, _DEFAULT)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24

    for ri in range(2, n_rows + 2):
        gen = gen_by_row[ri - 2] if gen_by_row and ri - 2 < len(gen_by_row) else None
        for ci, key in enumerate(col_keys, 1):
            cell = ws.cell(ri, ci)
            cell.border = BORDER
            cell.font = BODY_FONT
            _, center = COLUMN_STYLE.get(key, _DEFAULT)
            cell.alignment = CENTER if center else WRAP
            if gen and key in gen:
                cell.fill = AI_FILL  # AI 생성 셀 — 주황(지브라보다 우선)
            elif ri % 2 == 1:
                cell.fill = ZEBRA_FILL
        h = _row_height(ws, ri, col_keys)
        if h:
            ws.row_dimensions[ri].height = h

    ws.freeze_panes = "A2"
    if n_rows > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(col_keys))}{n_rows + 1}"


def style_summary_sheet(ws: Worksheet) -> None:
    """총괄표: 너비·제목 강조만 가볍게."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    if ws.max_row >= 1:
        ws.cell(1, 1).font = Font(bold=True, size=12)


# ── 개요 시트 (RFP 개요·요약·핵심기술·RISK) ──
_OV_TITLE_FONT = Font(bold=True, size=14, color="305496")
_OV_SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
_OV_SEC_FONT = Font(bold=True, size=11, color="1F3864")
_OV_HEAD_FONT = Font(bold=True, size=10)
_OV_ID_FILL = PatternFill("solid", fgColor="FCE4D6")
_RISK_LABEL = {"O": "가능", "△": "조건부", "X": "리스크"}


def write_overview_sheet(ws, requirements, recommendations, by_cat, category_spec, v2_overview=None):
    """RFP 개요 — 전체 요약 + 핵심 기술(보유/부족) + 핵심 RISK.

    v2_overview(LLM 생성: summary/techs/risks)가 있으면 그 서술형 요약·기술을 우선 사용.
    """
    from collections import Counter

    recs = recommendations or {}
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 26

    ws.cell(1, 1, "RFP 개요").font = _OV_TITLE_FONT
    row = 3

    def section(title: str):
        nonlocal row
        c = ws.cell(row, 1, title)
        c.fill = _OV_SEC_FILL
        c.font = _OV_SEC_FONT
        ws.cell(row, 2).fill = _OV_SEC_FILL
        ws.cell(row, 3).fill = _OV_SEC_FILL
        row += 1

    total = len(requirements)
    n_cat = len(by_cat)
    dist = Counter((recs.get(r.id).ai_risk if recs.get(r.id) else "") for r in requirements)
    if v2_overview and v2_overview.get("summary"):
        summary = str(v2_overview["summary"])
    else:
        summary = (
            f"총 {total}건의 요구사항을 {n_cat}개 분류로 정리했습니다. "
            f"AI 판정 — 가능(O) {dist.get('O', 0)} · 조건부(△) {dist.get('△', 0)} · "
            f"리스크(X) {dist.get('X', 0)} · 미산출 {dist.get('', 0)}. "
            f"분류 기준: {category_spec}"
        )
    section("전체 요약")
    cell = ws.cell(row, 1, summary)
    cell.alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.row_dimensions[row].height = 60
    row += 2

    # 핵심 기술 — KT 보유(채택 솔루션) / 부족 기술 집계
    section("핵심 기술")
    for ci, lbl in enumerate(("구분", "기술 / 솔루션", "관련 요구사항"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = _OV_HEAD_FONT
        c.fill = _OV_SEC_FILL
        c.border = BORDER
    row += 1

    owned: dict[str, list[str]] = {}
    missing: dict[str, list[str]] = {}
    for r in requirements:
        rec = recs.get(r.id)
        if not rec:
            continue
        for sku in getattr(rec, "matched_solution_skus", None) or []:
            label = sku.sku_label or sku.solution_name
            owned.setdefault(label, []).append(r.code)
        for t in getattr(rec, "missing_tech", None) or []:
            missing.setdefault(t, []).append(r.code)

    def tech_rows(label: str, data: dict[str, list[str]], fill_ids: bool):
        nonlocal row
        for name, ids in sorted(data.items(), key=lambda kv: -len(kv[1]))[:10]:
            ws.cell(row, 1, label)
            ws.cell(row, 2, name).alignment = WRAP
            idc = ws.cell(row, 3, ", ".join(ids[:8]) + (" 외" if len(ids) > 8 else ""))
            idc.alignment = WRAP
            if fill_ids:
                idc.fill = _OV_ID_FILL
            for ci in (1, 2, 3):
                ws.cell(row, ci).border = BORDER
                ws.cell(row, ci).font = BODY_FONT
            row += 1

    # RFP 핵심 기술 (V2 LLM 분석)
    v2_techs = (v2_overview or {}).get("techs") or []
    for item in v2_techs[:12]:
        name, req, ids = (list(item) + ["", "", ""])[:3]
        ws.cell(row, 1, "RFP 기술")
        ws.cell(row, 2, f"{name} — {req}" if req else str(name)).alignment = WRAP
        ws.cell(row, 3, str(ids)).alignment = WRAP
        ws.cell(row, 3).fill = _OV_ID_FILL
        for ci in (1, 2, 3):
            ws.cell(row, ci).border = BORDER
            ws.cell(row, ci).font = BODY_FONT
        row += 1
    if owned:
        tech_rows("KT 보유", owned, False)
    if missing:
        tech_rows("부족", missing, True)
    if not v2_techs and not owned and not missing:
        ws.cell(row, 1, "(기술 집계 없음)").font = Font(italic=True, size=10, color="808080")
        row += 1
    row += 1

    # 핵심 RISK — X(리스크) 판정 요구사항
    section("핵심 RISK (리스크 판정)")
    for ci, lbl in enumerate(("요구사항", "리스크 사유"), 1):
        c = ws.cell(row, ci, lbl)
        c.font = _OV_HEAD_FONT
        c.fill = _OV_SEC_FILL
        c.border = BORDER
    row += 1
    risks = [
        (r, recs.get(r.id))
        for r in requirements
        if recs.get(r.id) and recs.get(r.id).ai_risk == "X"
    ]
    if not risks:
        ws.cell(row, 1, "(리스크(X) 판정 요구사항 없음)").font = Font(
            italic=True, size=10, color="808080"
        )
        row += 1
    for r, rec in risks[:30]:
        idc = ws.cell(row, 1, f"{r.code} · {r.name}"[:60])
        idc.alignment = WRAP
        idc.fill = _OV_ID_FILL
        rc = ws.cell(row, 2, rec.ai_reason if rec else "")
        rc.alignment = WRAP
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2).border = BORDER
        row += 1

    ws.cell(1, 1).font = _OV_TITLE_FONT
