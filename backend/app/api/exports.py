from __future__ import annotations
import re

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel

from app.api.deps import ContainerDep
from app.domain.enums import ExportMode
from app.phase1.writers.export_columns import (
    EXPORT_COLUMNS,
    EXPORT_PRESETS,
    list_applicable_columns,
    resolve_export_columns,
)
from app.phase1.writers.sheet_writer import RequirementSheetWriter
from app.services.native_export import resolve_asset_path, write_native_excel

router = APIRouter(prefix="/documents", tags=["exports"])


_FIXED_COLS = ["요구사항 ID", "요구사항명", "계위", "상세내용",
               "KT 보유 기술", "부족 기술", "AI 판정", "AI 설명", "컨소시엄 필요 사항",
               "Human 판정", "Human 메모"]
_JUDG_MARK = {"YES": "O", "PARTIAL": "△", "NO": "X", "UNSET": ""}


def _expand_fixed_image_rows(group: list) -> list:
    """detail_images 가진 요건 뒤에 이미지 전용 행을 별도로 끼워넣는다(텍스트·이미지 비겹침).

    v2 excel_writer._expand_image_rows 와 동일한 컨벤션(표를 이미지로 렌더링해 셀에
    넣은 v_rule 표이미지 마커용) — 여기선 Requirement(pydantic)라 model_copy 로 복제.
    detail 이 "[표]" 플레이스홀더뿐(원문 마커 제거 후 남은 텍스트가 없던 경우, adapter.py
    참고)이면 그 자체를 별도 텍스트 행으로 내지 않고 첫 이미지 행의 캡션으로 붙인다 —
    안 그러면 "[표]"만 적힌 빈 행이 이미지 행과 분리되어 따로 남는다(실측 버그). 반환하는
    (row, imgs) 튜플의 row.detail 은 그 행에 실제로 찍을 텍스트로 이미 확정돼 있다."""
    out = []
    for r in group:
        imgs = r.detail_images
        if not imgs:
            out.append((r, []))
            continue
        detail = (r.detail or "").strip()
        is_caption_only = detail == "[표]" or not detail
        if not is_caption_only:
            out.append((r, []))          # 진짜 텍스트가 있으면 그건 그대로 한 행(이미지 없이)
        caption = detail if is_caption_only else "[표]"
        for i, p in enumerate(imgs):
            ir = r.model_copy(update={"detail": caption if i == 0 else "", "detail_images": [p]})
            out.append((ir, [p]))
    return out


def _fixed_excel(reqs, recs, judges, out_path, settings=None, doc_id: str = "", doc=None) -> None:
    """repo 요건 → 고정칼럼 조견표(탭=category). FE 편집이 즉시 반영(단일 소스=repo)."""
    import re as _re
    from collections import OrderedDict
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hdr = Font(bold=True, size=10, color="FFFFFF"); hfill = PatternFill("solid", fgColor="404040")
    ai_fill = PatternFill("solid", fgColor="FFF2CC"); wrap = Alignment(wrap_text=True, vertical="top")
    border = Border(*[Side(style="thin", color="D0D0D0")] * 4)
    by_tab: "OrderedDict[str, list]" = OrderedDict()
    for r in reqs:
        by_tab.setdefault(r.category or "요구사항", []).append(r)
    wb = openpyxl.Workbook(); wb.remove(wb.active); used: set[str] = set()
    for tab, group in by_tab.items():
        title = (_re.sub(r"[\\/*?:\[\]]", " ", tab).strip()[:31]) or "요구사항"
        b, i = title, 2
        while title in used:
            title = f"{b[:28]}_{i}"; i += 1
        used.add(title)
        ws = wb.create_sheet(title)
        for ci, h in enumerate(_FIXED_COLS, 1):
            c = ws.cell(1, ci, h); c.font = hdr; c.fill = hfill
        for ci, w in enumerate([16, 26, 20, 68, 20, 18, 8, 40, 24, 10, 24], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        rows = _expand_fixed_image_rows(group)
        for ri, (r, imgs) in enumerate(rows, 2):
            rec = recs.get(r.id); jud = judges.get(r.id)
            # r.detail 은 _expand_fixed_image_rows 가 이미 이 행에 찍을 텍스트로 확정해뒀다
            # (이미지 행의 첫 행엔 "[표]" 캡션, 나머지 이미지 행은 빈칸).
            vals = [r.code, r.name, (r.definition or ""), r.detail,
                    (rec.related_solution if rec else ""),
                    (", ".join(rec.missing_tech) if rec and rec.missing_tech else ""),
                    (_JUDG_MARK.get(rec.ai_risk.name, "") if rec else ""),
                    (rec.ai_reason if rec else ""),
                    (rec.consortium_need if rec and rec.consortium_need else ""),
                    (_JUDG_MARK.get(jud.mark.name, "") if jud else ""),
                    (jud.note if jud else "")]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v); c.alignment = wrap; c.border = border
                if ci >= 5:
                    c.fill = ai_fill
            if imgs and settings is not None:
                _embed_fixed_row_image(ws, ri, imgs[0], settings, doc_id, doc)
    if not by_tab:
        wb.create_sheet("요구사항")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _embed_fixed_row_image(ws, ri: int, rel_path: str, settings, doc_id: str, doc) -> None:
    """상세내용(4열) 셀에 표이미지 삽입 — 표를 인식해 HTML→PNG 렌더링한 것을 원문 대신 삽입."""
    path = resolve_asset_path(settings, doc_id, doc, rel_path)
    if path is None:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(path))
    except Exception:
        return
    scale = min(1.0, 640 / max(img.width, 1), 480 / max(img.height, 1))
    img.width = max(120, int(img.width * scale))
    img.height = max(80, int(img.height * scale))
    img.anchor = f"D{ri}"
    ws.add_image(img)
    ws.row_dimensions[ri].height = max(ws.row_dimensions[ri].height or 15, img.height * 0.8)


@router.get("/{doc_id}/export-fixed")
async def export_fixed(doc_id: str, container: ContainerDep):
    """고정칼럼(요구사항ID/명/계위/상세 + AI/Human) export — repo 직독이라 FE 편집 즉시 반영."""
    if doc_id not in container.repo.documents or not await container.repo.list_requirements(doc_id):
        await _restore_reqs_from_disk(container, doc_id)
    reqs, recs, judges = await container.repo.snapshot(doc_id)
    if not reqs:
        raise HTTPException(409, "추출된 요구사항 없음")
    out = container.settings.storage_root / doc_id / "exports" / "requirements_fixed.xlsx"
    doc = container.repo.documents.get(doc_id)
    _fixed_excel(reqs, recs, judges, out, settings=container.settings, doc_id=doc_id, doc=doc)
    fn = (getattr(doc, "display_name", None) or doc_id) + "_조견표.xlsx"
    return FileResponse(path=out, filename=fn,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Cache-Control": "no-store"})


async def _restore_reqs_from_disk(container, doc_id: str) -> bool:
    """repo 가 비었을 때(워크스페이스 reset·재시작) 디스크 v3_export.pkl 에서 문서+요건+추천 복원.

    추출은 storage_root/<doc>/v3_export.pkl 에 영속되므로, export 시 repo 에 없으면 되살린다.
    (다운로드가 reset/재시작에 견디게 — write_native_excel 은 pkl reqs 로 표를 만들고,
    복원된 app 요건+캐시 추천으로 AI 칼럼까지 병합). 반환: 복원 성공 여부.
    """
    import pickle
    from app.domain.enums import DocumentMime
    from app.domain.models import Document
    from app.services.artifact_cache import ArtifactCache
    from app.services.extraction import ExtractionService

    pkl = container.settings.storage_root / doc_id / "v3_export.pkl"
    if not pkl.is_file():
        return False
    try:
        payload = pickle.loads(pkl.read_bytes())
    except Exception:
        return False
    v2_reqs = payload.get("reqs") or []
    if not v2_reqs:
        return False
    # 문서 레코드가 없으면(재시작) 최소 placeholder 등록 — export 에 doc_id 만 있으면 충분
    if doc_id not in container.repo.documents:
        incoming = container.settings.storage_root / "incoming"
        src = next(iter(incoming.glob(f"{doc_id}.*")), pkl)
        await container.repo.save_document(Document(id=doc_id, src_path=src, mime=DocumentMime.PDF))
    app_reqs = [ExtractionService._v2_to_requirement(doc_id, r) for r in v2_reqs]
    await container.repo.save_requirements(doc_id, app_reqs)
    doc = container.repo.documents.get(doc_id)
    if doc is not None:
        try:
            await ArtifactCache(container.settings.artifact_cache_dir).restore_recommendations(
                container, doc, fast=True
            )
        except Exception:
            pass
    return True


class ExportColumnInfo(BaseModel):
    key: str
    header: str
    group: str


class ExportColumnsResponse(BaseModel):
    preset: str
    mode: str
    selected: list[str]
    applicable: list[ExportColumnInfo]
    presets: dict[str, list[str]]


@router.get("/{doc_id}/export/columns", response_model=ExportColumnsResponse)
async def list_export_columns(
    doc_id: str,
    container: ContainerDep,
    mode: ExportMode = ExportMode.BOTH,
    preset: str = "standard",
) -> ExportColumnsResponse:
    if doc_id not in container.repo.documents:
        raise HTTPException(404, f"document 없음: {doc_id}")
    reqs, recs, judges = await container.repo.snapshot(doc_id)
    preset_key = preset if preset in EXPORT_PRESETS else "조견표"
    selected = resolve_export_columns(reqs, recs, judges, mode, EXPORT_PRESETS[preset_key])
    applicable_keys = list_applicable_columns(reqs, recs, judges, mode)
    return ExportColumnsResponse(
        preset=preset_key,
        mode=mode.value,
        selected=selected,
        applicable=[
            ExportColumnInfo(key=k, header=EXPORT_COLUMNS[k].header, group=EXPORT_COLUMNS[k].group)
            for k in applicable_keys
        ],
        presets=EXPORT_PRESETS,
    )


@router.get("/{doc_id}/export")
async def export_excel(
    doc_id: str,
    container: ContainerDep,
    mode: ExportMode = ExportMode.BOTH,
    cols: str | None = None,
    adaptive: bool = True,
    layout: str = "ordered",
    filename: str | None = None,
) -> FileResponse:
    # repo 가 비었으면(reset·재시작·추천 중 reset) 디스크 v3_export.pkl 에서 문서+요건 복원
    if doc_id not in container.repo.documents or not await container.repo.list_requirements(doc_id):
        await _restore_reqs_from_disk(container, doc_id)
    if doc_id not in container.repo.documents:
        raise HTTPException(404, f"document 없음: {doc_id}")
    reqs, recs, judges = await container.repo.snapshot(doc_id)
    if not reqs:
        raise HTTPException(409, "추출된 요구사항 없음")

    doc = container.repo.documents.get(doc_id)
    out_dir = container.settings.storage_root / doc_id / "exports"
    out_dir.mkdir(parents=True, exist_ok=True)

    # V3/V2 prototype 조견표 — 표안표·이미지 포함 원문순서 Excel (백엔드 파이프라인과 동일)
    native_path = out_dir / "requirements_native.xlsx"
    if write_native_excel(
        container.settings,
        doc_id,
        doc,
        native_path,
        app_reqs=reqs,
        recommendations=recs,
        judgements=judges,
    ):
        return FileResponse(
            path=native_path,
            filename=_download_name(filename, native_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Cache-Control": "no-store"},
        )

    layout_key = "ordered" if layout == "ordered" else "cluster"
    column_keys = [c.strip() for c in cols.split(",")] if cols else None

    col_tag = cols.replace(",", "-")[:40] if cols else "adaptive"
    out_path = out_dir / f"requirements_{mode.value}_{layout_key}_{col_tag}.xlsx"

    # V2 LLM 개요(요약·핵심기술·리스크)가 있으면 개요 시트에 사용 (AI 칼럼은 앱 writer가 포함)
    v2_overview = _load_v2_overview(container, doc_id)
    RequirementSheetWriter().write(
        out_path,
        reqs,
        recommendations=recs,
        judgements=judges,
        mode=mode,
        columns=column_keys,
        adaptive=adaptive,
        layout=layout_key,
        v2_overview=v2_overview,
        extraction_meta=container.repo.documents.get(doc_id).extraction_meta
        if doc_id in container.repo.documents
        else None,
    )
    return FileResponse(
        path=out_path,
        filename=_download_name(filename, out_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Cache-Control": "no-store"},
    )


def _download_name(filename: str | None, out_path) -> str:
    """사용자 지정 다운로드 파일명 (없으면 out_path 이름). 경로 문자 제거."""
    if filename:
        safe = re.sub(r'[\\/:*?"<>|]+', "_", filename).strip() or out_path.stem
        return safe if safe.lower().endswith(".xlsx") else f"{safe}.xlsx"
    return out_path.name


def _load_v2_overview(container, doc_id: str):
    """V2 추출 시 생성한 LLM 개요(summary/techs/risks) dict — 없으면 None."""
    import pickle
    from pathlib import Path

    candidates = [
        container.settings.storage_root / doc_id / "v3_export.pkl",
        container.settings.storage_root / doc_id / "v2_export.pkl",
    ]
    doc = container.repo.documents.get(doc_id)
    if doc is not None and getattr(doc, "content_hash", None):
        bucket = container.settings.artifact_cache_dir / doc.content_hash[:16]
        candidates.extend([bucket / "v3_export.pkl", bucket / "v2_export.pkl"])
    pkl = next((p for p in candidates if Path(p).is_file()), None)
    if pkl is None:
        return None
    try:
        payload = pickle.loads(Path(pkl).read_bytes())
        return payload.get("overview")
    except Exception:  # noqa: BLE001
        return None
